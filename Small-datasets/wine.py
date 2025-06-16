import torch
import torch.nn as nn
import torch.optim as optim
from sklearn.datasets import load_wine
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import numpy as np
import random
import argparse
import time
import os 
# --- Argument Parsing ---
parser = argparse.ArgumentParser(description='Run CNO or CNOGNP optimization on Breast Cancer dataset.')

parser.add_argument('--algorithm', type=str, default='CNO', choices=['CNO', 'CNOGNP', 'PSO', 'PSO-BFGS'],
                    help='Optimization algorithm to use: CNO or CNOGNP (default: CNO)')
parser.add_argument('--n_particles', type=int, default=50,
                    help='Number of particles (swarm size) (default: 50)')
parser.add_argument('--max_iterations', type=int, default=500,
                    help='Maximum number of iterations (default: 500)')
parser.add_argument('--omega', type=float, default=0.9,
                    help='Inertia weight (default: 0.9)')
parser.add_argument('--eta', type=float, default=0.01,
                    help='Scale factor for gradient step (default: 0.01)')
parser.add_argument('--c1', type=float, default=2.0,
                    help='Cognitive learning factor (default: 2.0)')
parser.add_argument('--c2', type=float, default=2.0,
                    help='Social learning factor (default: 2.0)')
parser.add_argument('--lambda_reg', type=float, default=0.01,
                    help='Regularization parameter for gradient norm in CNOGNP (default: 0.01). Only used if --algorithm is CNOGNP.')
parser.add_argument("--row_num", type=int, default=20, help="row_num (currently unused)") # Kept as requested
parser.add_argument("--column_num", type=int, default=20, help="column_num (currently unused)") # Kept as requested
parser.add_argument("--hidden_dim", type=int, default=10, help="Number of neurons in hidden layer (default: 10)") # Used in SimpleMLP
parser.add_argument("--loc", type=str, default="./results", help="Directory to save results and plots (default: ./results)")


args = parser.parse_args()


if torch.cuda.is_available():
    device = torch.device("cuda")
    print("CUDA is available! Using GPU.")
else:
    device = torch.device("cpu")
    print("CUDA not available. Using CPU.")



# 1. Load and preprocess the breast cancer dataset
data = load_wine()
X, y = data.data, data.target

num_classes = len(np.unique(y))
print(f"Number of classes: {num_classes}")


# Split data
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2) 

# Scale features
scaler = StandardScaler()
X_train_scaled = scaler.fit_transform(X_train)
X_test_scaled = scaler.transform(X_test)

# Convert to PyTorch tensors
X_train_tensor = torch.tensor(X_train_scaled, dtype=torch.float32).to(device) # Add .to(device)
X_test_tensor = torch.tensor(X_test_scaled, dtype=torch.float32).to(device) # Add .to(device)
# Change dtype to long and remove unsqueeze(1)
y_train_tensor = torch.tensor(y_train, dtype=torch.long).to(device)
y_test_tensor = torch.tensor(y_test, dtype=torch.long).to(device)
# Get input dimension
input_dim = X_train_tensor.shape[1]

# 2. Define a simple MLP model
class SimpleMLP(nn.Module):
    def __init__(self, input_dim, hidden_dim, output_dim):
        super(SimpleMLP, self).__init__()
        self.layer1 = nn.Linear(input_dim, hidden_dim)
        self.relu = nn.ReLU()
        self.layer2 = nn.Linear(hidden_dim, output_dim)
        # Using BCEWithLogitsLoss later, so no sigmoid here

        # Initialize weights (optional, but often good practice)
        nn.init.xavier_uniform_(self.layer1.weight)
        nn.init.zeros_(self.layer1.bias)
        nn.init.xavier_uniform_(self.layer2.weight)
        nn.init.zeros_(self.layer2.bias)

    def forward(self, x):
        x = self.layer1(x)
        x = self.relu(x)
        x = self.layer2(x)
        return x

    # --- Moved from BaseParticle ---
    def _get_parameters_vector(self):
        """Helper to flatten model parameters into a single vector."""
        return torch.cat([p.view(-1) for p in self.parameters()])

    def _set_parameters_from_vector(self, param_vector):
        """Helper to set model parameters from a single vector."""
        offset = 0
        # Ensure param_vector is on the same device as model parameters if using GPU
        param_vector = param_vector.to(next(self.parameters()).device)
        for p in self.parameters():
            num_elements = p.numel()
            # Ensure the slice has the correct shape before copying
            with torch.no_grad(): # Avoid tracking gradients during parameter update
                 p.data.copy_(param_vector[offset : offset + num_elements].view(p.size()))
            offset += num_elements
    # --- End Moved Methods ---


# Loss function 
criterion = nn.CrossEntropyLoss()

# 3. Define a base Particle class
class BaseParticle:
    def __init__(self, model_template, device):
        """
        Initializes a base particle.

        Args:
            model_template: A function that returns a new instance of the MLP model.
        """
        # Create the model instance - this handles random weight initialization
        # Use the hidden_dim from args
        self.device = device # Store device
        self.model = model_template(input_dim=input_dim, hidden_dim=args.hidden_dim, output_dim=num_classes).to(self.device) # Add .to(self.device)
        # Initialize core particle attributes (position, velocity, personal best position)
        self._initialize_core_attributes()

        # Training loss and fitness value (initialized to inf)
        self.training_loss = float('inf') # Store the most recently calculated training loss
        self.fitness_value = float('inf')
        self.personal_best_fitness = float('inf') # Fitness at personal best position

    def _initialize_core_attributes(self):
        """
        Initializes the particle's position, velocity, and personal best position.
        Corresponds to algorithm steps 1, 2, and 3.
        This method is called once during __init__.
        """
        # Particle position: model parameters flattened into a single vector (Step 1)
        # Call the method on the model instance
        self.position = self.model._get_parameters_vector()

        # Particle velocity (initialized to zero) (Step 2)
        self.velocity = torch.zeros_like(self.position).to(self.device) # Ensure velocity starts on device
        # Personal best position (initialized to current position) (Step 3)
        self.personal_best_position = self.position.clone().to(self.device) # Ensure p_best starts on device


    # Removed _get_parameters_vector and _set_parameters_from_vector from here

    def particle_forward_backward(self, X, y):
        """
        Calculate the gradient vector of the training loss f(z^i) with respect to parameters.
        This is used for the gradient step (Line 7). It does NOT update self.training_loss
        or store gradients in self.model.parameters().grad permanently for later use,
        as the position changes after this step.
        """
        # Call the method on the model instance
        self.model._set_parameters_from_vector(self.position) # Ensure model parameters match position
        self.model.train() # Set model to training mode
        self.model.zero_grad() # Clear previous gradients

        outputs = self.model(X)
        loss = criterion(outputs, y) # Use the global criterion

        # Compute gradients
        loss.backward()
        self.training_loss = loss.item()

        # Flatten gradients
        grad_vector = torch.cat([p.grad.view(-1) if p.grad is not None else torch.zeros_like(p.view(-1))
                                 for p in self.model.parameters()])

        return grad_vector
    
    def particle_forward(self, X, y):
        """
        Calculate the gradient vector of the training loss f(z^i) with respect to parameters.
        This is used for the gradient step (Line 7). It does NOT update self.training_loss
        or store gradients in self.model.parameters().grad permanently for later use,
        as the position changes after this step.
        """
        # Call the method on the model instance
        self.model._set_parameters_from_vector(self.position) # Ensure model parameters match position
        self.model.train() # Set model to training mode
        self.model.zero_grad() # Clear previous gradients

        outputs = self.model(X)
        loss = criterion(outputs, y) # Use the global criterion

        # Compute gradients
        
        self.training_loss = loss.item()
 

    # def calculate_loss_and_populate_gradients(self, X, y):
    #     """
    #     Performs a forward/backward pass at the *current* position (after Line 9 update).
    #     Updates self.training_loss and populates self.model.parameters().grad.
    #     This is done *before* calculate_fitness is called in the step method.
    #     """
    #     # Call the method on the model instance
    #     self.model._set_parameters_from_vector(self.position) # Ensure model parameters match position
    #     self.model.train() # Set model to training mode
    #     self.model.zero_grad() # Clear previous gradients

    #     outputs = self.model(X)
    #     loss = criterion(outputs, y) # Use the global criterion

    #     # Compute gradients - this populates self.model.parameters().grad
    #     loss.backward()

    #     # Update the training_loss attribute with the loss calculated at the current position
    #     self.training_loss = loss.item()


    def calculate_fitness(self):
        """
        Abstract method to calculate fitness. Must be implemented by subclasses.
        This method now assumes self.training_loss and self.model.parameters().grad
        are already populated for the current position by calculate_loss_and_populate_gradients.
        """
        raise NotImplementedError("Subclass must implement abstract method")

    def update_personal_best(self):
        """Update personal best position and fitness if current fitness is better."""
        # Assuming lower fitness is better
        if self.fitness_value < self.personal_best_fitness:
            self.personal_best_fitness = self.fitness_value
            self.personal_best_position = self.position.clone()

    def step(self, global_best_position, omega, eta, c1, c2, X_train, y_train):
        """Perform one optimization step (lines 7-16 from the algorithm)."""
        # Line 7: Gradient step - calculate gradient at the *current* position (before velocity/position update)
        nabla_f = self.particle_forward_backward(X_train, y_train)

        z_bar_i = self.position - eta * nabla_f

        # Line 8: PSO velocity update (uses current position self.position)
        r1, r2 = torch.rand(2) # Random numbers for cognitive and social components
        cognitive_component = c1 * r1 * (self.personal_best_position - self.position)
        social_component = c2 * r2 * (global_best_position - self.position)
        self.velocity = omega * (z_bar_i - self.position) + cognitive_component + social_component

        # Line 9: Position update (gradient step result + velocity)
        self.position = self.position + self.velocity

        # Update model parameters from the new position
        # Call the method on the model instance
        self.model._set_parameters_from_vector(self.position)


        # --- END CRUCIAL STEP ---


        # Line 10: Evaluate fitness at the *new* position
        # This call now uses the self.training_loss and self.model.parameters().grad
        # that were just populated by calculate_loss_and_populate_gradients.
        self.calculate_fitness() # No X, y needed here anymore

        # Lines 11-13: Update personal best
        self.update_personal_best()

        # Lines 14-16: Global best update is handled outside this method
        # in the main loop, comparing self.fitness_value to global_best_fitness


# 4. Define ParticleA (CNO)
class ParticleA(BaseParticle):
    def __init__(self, model_template, device):
        super().__init__(model_template, device)

    def calculate_fitness(self):
        """Fitness for CNO is the training loss f(z^i)."""
        # For CNO, phi(z^i) = f(z^i)
        # The training_loss attribute was updated by calculate_loss_and_populate_gradients
        # in the step method, so we can just use it directly.
        self.fitness_value = self.training_loss
        return self.fitness_value

# 5. Define ParticleB (CNOGNP)
class ParticleB(BaseParticle):
    def __init__(self, model_template, lambda_reg, device):
        """
        Initializes a ParticleB (CNOGNP).

        Args:
            model_template: A function that returns a new instance of the MLP model.
            lambda_reg: Regularization parameter for the gradient norm term.
        """
        super().__init__(model_template, device)
        self.lambda_reg = lambda_reg # Regularization parameter for gradient norm

    def calculate_fitness(self):
        """Fitness for CNOGNP is training loss f(z^i) + lambda * ||nabla f(z^i)||_2."""
        # For CNOGNP, phi(z^i) = f(z^i) + lambda * ||nabla f(z^i)||_2
        # The training_loss attribute and model.parameters().grad were populated
        # by calculate_loss_and_populate_gradients in the step method.

        training_loss = self.training_loss

        # Calculate L2 norm of the gradients that are already in self.model.parameters().grad
        grad_norm = 0
        for p in self.model.parameters():
            if p.grad is not None:
                grad_norm += p.grad.norm(2).item()**2 # Sum of squared norms
        grad_norm = np.sqrt(grad_norm) # Take square root for the total L2 norm


        # Calculate fitness using the pre-calculated loss and gradient norm
        self.fitness_value = training_loss + self.lambda_reg * grad_norm
        return self.fitness_value

# 6. Define ParticlePSO (Standard PSO)
class ParticlePSO(BaseParticle):
    def __init__(self, model_template, device):
        """
        Initializes a ParticlePSO (Standard PSO).
        """
        super().__init__(model_template, device)
        # No extra parameters needed for standard PSO compared to BaseParticle

    def calculate_fitness(self):
        """Fitness for standard PSO is the training loss f(z^i)."""
        # Fitness is just the training loss, which should be updated
        # by particle_forward called within the step method.
        self.fitness_value = self.training_loss
        return self.fitness_value
    

    # Override the step method for standard PSO logic
    def step(self, global_best_position, omega, eta, c1, c2, X_train, y_train):
        """Perform one standard PSO optimization step."""
        # NOTE: Standard PSO does NOT use the gradient step (eta * nabla_f)
        # like CNO/CNOGNP. It updates velocity based on current velocity,
        # personal best, and global best.

        # PSO velocity update (Standard formula)
        r1, r2 = torch.rand(2, device=self.device) # Random numbers on the correct device
        cognitive_component = c1 * r1 * (self.personal_best_position - self.position)
        social_component = c2 * r2 * (global_best_position - self.position)
        # Use self.velocity (previous velocity)
        self.velocity = omega * self.velocity + cognitive_component + social_component

        # Position update (Standard formula)
        self.position = self.position + self.velocity

        # Update model parameters from the new position
        self.model._set_parameters_from_vector(self.position)

        # --- CRUCIAL STEP for PSO/CNO/CNOGNP ---
        # Evaluate loss at the *new* position to determine fitness and update p_best
        # We call particle_forward which calculates loss and updates self.training_loss
        # We don't need the returned gradient vector here, but calculating it is part of particle_forward
        _ = self.particle_forward(X_train, y_train) # Updates self.training_loss
        # --- END CRUCIAL STEP ---

        # Calculate fitness at the *new* position using the updated self.training_loss
        self.calculate_fitness() # Uses the self.training_loss updated above

        # Update personal best based on the new fitness
        self.update_personal_best()

        # Global best update is handled outside this method
    
# 7. Define ParticleBFGS (PSO + L-BFGS)
class ParticleBFGS(BaseParticle):
    def __init__(self, model_template, device, eta):
        """
        Initializes a ParticleBFGS (PSO + L-BFGS).

        Args:
            model_template: A function that returns a new instance of the MLP model.
            device: The torch device (cpu or cuda).
            eta: The learning rate for the L-BFGS optimizer.
        """
        super().__init__(model_template, device)
        # The L-BFGS optimizer is stateful, so each particle needs its own instance.
        # It operates directly on the particle's model parameters.
        # The `eta` parameter from CNO serves as the learning rate here.
        self.optimizer = optim.LBFGS(self.model.parameters(), lr=eta, max_iter=5) # max_iter can be tuned

    def calculate_fitness(self):
        """Fitness for PSO-BFGS is the training loss f(z^i), same as CNO."""
        # The training_loss attribute is updated after the full particle step.
        self.fitness_value = self.training_loss
        return self.fitness_value

    # Override the step method for the specific PSO-BFGS logic
    def step(self, global_best_position, omega, eta, c1, c2, X_train, y_train):
        """Perform one PSO-BFGS optimization step."""
        # Ensure the model's parameters match the particle's current position
        self.model._set_parameters_from_vector(self.position)

        # --- Line 7 (Modified): BFGS Step ---
        # The L-BFGS optimizer in PyTorch requires a "closure" function that
        # re-evaluates the model and returns the loss.
        def closure():
            self.optimizer.zero_grad()
            outputs = self.model(X_train)
            loss = criterion(outputs, y_train)
            loss.backward()
            return loss

        # The optimizer.step() call performs the BFGS update(s) on self.model.parameters
        self.optimizer.step(closure)

        # After the BFGS step, the model's parameters have been updated.
        # This new parameter vector is our z_bar_i.
        z_bar_i = self.model._get_parameters_vector()
        # --- End of BFGS Step ---

        # --- Lines 8 & 9: Standard CNO Velocity and Position Update ---
        # The rest of the update follows the CNO structure, using z_bar_i.
        r1, r2 = torch.rand(2, device=self.device)
        cognitive_component = c1 * r1 * (self.personal_best_position - self.position)
        social_component = c2 * r2 * (global_best_position - self.position)

        # The velocity update uses the difference between the BFGS-updated position
        # and the original position.
        self.velocity = omega * (z_bar_i - self.position) + cognitive_component + social_component

        # The final position update is the original position plus the new velocity.
        # Note: We use self.position here, NOT z_bar_i.
        self.position = self.position + self.velocity

        # Update the model's parameters to reflect the final new position
        self.model._set_parameters_from_vector(self.position)

        # --- Lines 10-16: Evaluation and Updates ---
        # Evaluate loss at the *final new* position to determine fitness
        self.particle_forward(X_train, y_train) # This updates self.training_loss

        # Calculate fitness based on the new loss
        self.calculate_fitness()

        # Update personal best
        self.update_personal_best()

        # Global best update is handled outside this method in the main loop.

# Add test evaluation function
def evaluate_model(model, X_test, y_test, device):
    """Evaluates the model on the test set."""
    model.eval() # Set model to evaluation mode
    model.to(device)
    X_test = X_test.to(device) # Ensure test data is on the correct device
    y_test = y_test.to(device) # Ensure test labels are on the correct device
    with torch.no_grad(): # No gradient calculation needed for evaluation
        outputs = model(X_test)
        loss = criterion(outputs, y_test)

        # Calculate accuracy for multi-class
        _, predicted = torch.max(outputs, 1) # Get the index of the max log-probability
        correct = (predicted == y_test).sum().item()
        accuracy = correct / y_test.size(0)

    return loss.item(), accuracy

# Main Optimization Algorithm Function
def run_optimization(algorithm_type, N, kappa_max, omega, eta, c1, c2, lambda_reg=None, device=None):
    """
    Runs the specified optimization algorithm (CNO or CNOGNP).

    Args:
        algorithm_type (str): 'CNO' or 'CNOGNP'.
        N (int): Particle number.
        kappa_max (int): Maximum number of iterations.
        omega (float): Inertia weight.
        eta (float): Scale factor for gradient step.
        c1 (float): Cognitive learning factor.
        c2 (float): Social learning factor.
        lambda_reg (float, optional): Regularization parameter for CNOGNP.

    Returns:
        tuple: (best_model, g_best_fitness, test_loss, test_accuracy)
    """
    print(f"\n--- Running {algorithm_type} algorithm ---")
    print(f"Parameters: N={N}, kappa_max={kappa_max}, omega={omega}, eta={eta}, c1={c1}, c2={c2}")
    if algorithm_type == 'CNOGNP':
        print(f"CNOGNP specific: lambda_reg={lambda_reg}")

    # Create the swarm
    particles = []
    # Define the model template function here to use the hidden_dim from args
    def create_mlp_model_instance(input_dim, hidden_dim, output_dim): # <-- Add output_dim here
        return SimpleMLP(input_dim=input_dim, hidden_dim=hidden_dim, output_dim=output_dim) # <-

    for _ in range(N):
        if algorithm_type == 'CNO':
            particles.append(ParticleA(model_template=create_mlp_model_instance, device=device))
        elif algorithm_type == 'CNOGNP':
             if lambda_reg is None:
                 # This check is also done by argparse default, but good to be explicit
                 raise ValueError("lambda_reg must be provided for CNOGNP")
             particles.append(ParticleB(model_template=create_mlp_model_instance, lambda_reg=lambda_reg, device=device))
        elif algorithm_type == 'PSO': # <-- ADD THIS BLOCK
             particles.append(ParticlePSO(model_template=create_mlp_model_instance, device=device)) # <-- ADD THIS LINE
            # --- ADD THIS NEW BLOCK ---
        elif algorithm_type == 'PSO-BFGS':
             # The eta parameter is used as the learning rate for the LBFGS optimizer
             particles.append(ParticleBFGS(model_template=create_mlp_model_instance, device=device, eta=eta))
        # --- END OF NEW BLOCK ---
        else:
            # This should not happen due to argparse choices, but as a safeguard
            raise ValueError("Invalid algorithm_type. Choose 'CNO' or 'CNOGNP'.")

    # Initialize global best
    g_best_fitness = float('inf')
    g_best_position = None # Will store the parameter vector

    # Evaluate initial fitness and set initial personal/global bests
    print("Initializing swarm and global best...")
    for i, particle in enumerate(particles):
        # For initial fitness, we need to calculate loss and gradients at the starting position.
        # Call the method that does forward/backward and populates attributes.
        particle.particle_forward_backward(X_train_tensor, y_train_tensor)

        # Now calculate fitness using the populated attributes
        particle.calculate_fitness() # No X, y needed here

        # Update personal best (initial position is the first personal best)
        particle.update_personal_best()

        # Update global best if this particle is the best so far
        if particle.fitness_value < g_best_fitness:
            g_best_fitness = particle.fitness_value
            g_best_position = particle.position.clone()
            # print(f"Initial Global Best Fitness updated by particle {i}: {g_best_fitness:.4f}") # Optional print

    print(f"Initial Global Best Fitness: {g_best_fitness:.4f}")

    # Main optimization loop
    print("Starting optimization loop...")
    for kappa in range(kappa_max):
        # print(f"Iteration {kappa+1}/{kappa_max}") # Optional print per iteration
        for i, particle in enumerate(particles):
            # Perform the particle step (lines 7-16 from the algorithm)
            # The step method now includes the forward/backward pass at the new position
            # before calling calculate_fitness.
            particle.step(g_best_position, omega, eta, c1, c2, X_train_tensor, y_train_tensor)

            # Update global best (Lines 14-16)
            # Note: The fitness is calculated *after* the position update in particle.step
            if particle.fitness_value < g_best_fitness:
                g_best_fitness = particle.fitness_value
                g_best_position = particle.position.clone()
                # print(f"Global Best Fitness updated at iteration {kappa+1} by particle {i}: {g_best_fitness:.4f}") # Optional print

        # Optional: Print global best fitness periodically
        if (kappa + 1) % 100 == 0 or kappa == 0 or kappa == kappa_max - 1:
             print(f"Iteration {kappa+1}: Global Best Fitness = {g_best_fitness:.4f}")


    print("\nOptimization finished.")
    print(f"Final Global Best Fitness: {g_best_fitness:.4f}")

    # Evaluate the global best model on the test set
    print("Evaluating the global best model on the test set...")
    # Create a new model instance for the final evaluation
    # Use the hidden_dim from args
    best_model = SimpleMLP(input_dim=input_dim, hidden_dim=args.hidden_dim, output_dim=num_classes).to(device)
    # Set its parameters to the global best position found - NOW CALLING ON SimpleMLP INSTANCE
    best_model._set_parameters_from_vector(g_best_position)

    test_loss, test_accuracy = evaluate_model(best_model, X_test_tensor, y_test_tensor, device)

    print(f"Test Loss: {test_loss:.4f}")
    print(f"Test Accuracy: {test_accuracy:.4f}")

    return best_model, g_best_fitness, test_loss, test_accuracy

# --- Run the selected algorithm using parsed arguments ---
if __name__ == "__main__":
    # Check if lambda_reg is needed for CNOGNP
    if args.algorithm == 'CNOGNP':
        if args.lambda_reg is None:
             # This case is handled by argparse default, but being explicit
             print("Warning: lambda_reg not specified for CNOGNP, using default 0.01")
        lambda_val = args.lambda_reg
    else:
        lambda_val = None # lambda_reg is not used for CNO

    start_time = time.perf_counter()

    best_model, g_best_fitness, test_loss, test_accuracy = run_optimization(
        algorithm_type=args.algorithm,
        N=args.n_particles,
        kappa_max=args.max_iterations,
        omega=args.omega,
        eta=args.eta,
        c1=args.c1,
        c2=args.c2,
        lambda_reg=lambda_val,
        device=device
    )

    end_time = time.perf_counter()

    # 4. Calculate the elapsed time
    elapsed_time = end_time - start_time

    print("\n" + "="*30)
    print(f"Final Results for {args.algorithm} Algorithm")
    print("="*30)
    print(f"Final Global Best Fitness: {g_best_fitness:.4f}")
    print(f"Test Loss: {test_loss:.4f}")
    print(f"Test Accuracy: {test_accuracy:.4f}")
    print(f"--- Execution Time: {elapsed_time:.4f} seconds ---") # Print the duration



    from openpyxl import load_workbook
    while True: # <-- Added: Start retry loop
        try:
            excel_dir='log/wine/{}/{}.xlsx'.format(args.loc,args.loc) #正常对比实验
            wb = load_workbook(excel_dir)
            sheet = wb.active
            row_num=args.row_num
            column_num=args.column_num+2
            if args.algorithm == 'PSO':
                label = f'PSO_I{args.max_iterations}_P{args.n_particles}_H{args.hidden_dim}_W{args.omega}_C1{args.c1}_C2{args.c2}' # <-- ADD THIS LINE
            if args.algorithm == 'CNO':
                label = f'CNO_I{args.max_iterations}_P{args.n_particles}_H{args.hidden_dim}_E{args.eta}_W{args.omega}_C1{args.c1}_C2{args.c2}'
            if args.algorithm == 'CNOGNP':
                label = f'CNOGNP_Opt_L{args.lambda_reg}_I{args.max_iterations}_P{args.n_particles}_H{args.hidden_dim}_E{args.eta}_W{args.omega}_C1{args.c1}_C2{args.c2}'
            # --- ADD THIS NEW BLOCK ---
            if args.algorithm == 'PSO-BFGS':
                label = f'PSO-BFGS_I{args.max_iterations}_P{args.n_particles}_H{args.hidden_dim}_E{args.eta}_W{args.omega}_C1{args.c1}_C2{args.c2}'
            # --- END OF NEW BLOCK ---
            sheet.cell(row=row_num, column=1).value = label
            sheet.cell(row=row_num, column=column_num).value = test_accuracy
            sheet.cell(row=row_num, column=column_num+12).value = elapsed_time # Log time

            wb.save(excel_dir)
            break # <-- Added: Exit loop on success
        except PermissionError: # <-- Added: Catch file access error
            # If file is locked, wait for 1 second and retry
            print(f"Excel file '{excel_dir}' is locked, retrying in 1 second...")
            time.sleep(1) # <-- Added: Wait before retrying
    print("运行结束！")



